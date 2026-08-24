"""Builds Excel workbooks from scan/cleanup results: the full file-wise
analysis, a folder-scoped export, and a Cleanup Center export - all sharing
the same "Action Options" dropdown pattern."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from theme import ACCENT_TEXT_ON_LIGHT, CARD_BG, TEXT_ON_DARK, human_size

HEADER_FILL = PatternFill(start_color=f"FF{CARD_BG.lstrip('#')}", end_color=f"FF{CARD_BG.lstrip('#')}", fill_type="solid")
HEADER_FONT = Font(color=f"FF{TEXT_ON_DARK.lstrip('#')}", bold=True)
HYPERLINK_FONT = Font(color=f"FF{ACCENT_TEXT_ON_LIGHT.lstrip('#')}", underline="single")

# Seed values for the editable dropdown list. The user can add/remove rows
# on the "Action Options" sheet directly - the dropdown re-reads that range,
# it isn't a fixed/hardcoded list baked into the validation.
DEFAULT_ACTIONS = ["Delete", "Keep", "Review", "Archive", "Ignore"]
ACTION_LIST_ROWS = 30  # how many rows of the Action Options sheet feed the dropdown


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_action_options_sheet(wb):
    """Adds the editable 'Action Options' sheet plus a workbook-level named
    range ('ActionList') pointing at it. A named range - rather than a raw
    quoted cross-sheet reference ('Action Options'!$A$2:$A$31 written
    directly into a data validation's formula1) - is what actually makes the
    dropdown populate: Excel's desktop client reliably renders an EMPTY
    in-cell list for list-validations that reference another sheet directly,
    even though the XML is otherwise valid. A named range is the standard,
    broadly-compatible fix (works the same in Excel, LibreOffice, and Google
    Sheets) and still keeps the list itself user-editable on this sheet.
    """
    opts = wb.create_sheet("Action Options")
    opts.cell(row=1, column=1, value="Action").font = Font(bold=True)
    for i, val in enumerate(DEFAULT_ACTIONS, start=2):
        opts.cell(row=i, column=1, value=val)
    opts.column_dimensions["A"].width = 24
    opts.cell(row=ACTION_LIST_ROWS + 3, column=1,
              value="Add your own actions above (within the highlighted range) - the dropdown updates automatically.")
    opts.cell(row=ACTION_LIST_ROWS + 3, column=1).font = Font(italic=True, color="FF898781")

    wb.defined_names["ActionList"] = DefinedName(
        "ActionList", attr_text=f"'Action Options'!$A$2:$A${ACTION_LIST_ROWS + 1}"
    )


def add_action_dropdown(ws, col_letter, n_rows):
    """Attaches the Action dropdown (sourced from the ActionList named range -
    see add_action_options_sheet, which must already have been called on this
    workbook) to `col_letter` for rows 2..n_rows."""
    if n_rows < 2:
        return
    dv = DataValidation(
        type="list",
        formula1="ActionList",
        allow_blank=True,
        showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk: this flag is inverted)
    )
    dv.error = "Pick a value from the list, or edit the 'Action Options' sheet to add your own."
    dv.errorTitle = "Not in the list"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{n_rows}")


def _split_path(path):
    folder = path.rsplit("\\", 1)[0] if "\\" in path else ""
    name = path.rsplit("\\", 1)[-1]
    ext = ("." + name.rsplit(".", 1)[-1]) if "." in name else ""
    return folder, name, ext


def build_excel(data, safe_files, out_path):
    wb = Workbook()
    wb.active.title = "Summary"  # placeholder; filled in below, kept first for readability

    # ---- Sheet: file-wise, safe-to-delete list ----
    ws = wb.create_sheet("Safe to Delete - Files")
    headers = ["Path", "Folder", "File name", "Category", "Extension", "Size (bytes)", "Size", "Action", "Open"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for f in safe_files:
        path = f["path"]
        folder, name, ext = _split_path(path)
        row = [path, folder, name, f["category"], ext, f["size"], human_size(f["size"]), "", "Open"]
        ws.append(row)
        r = ws.max_row
        open_cell = ws.cell(row=r, column=9)
        open_cell.hyperlink = path
        open_cell.font = HYPERLINK_FONT
        open_cell.alignment = Alignment(horizontal="center")

    n_rows = len(safe_files) + 1
    add_action_options_sheet(wb)
    if n_rows > 1:
        add_action_dropdown(ws, "H", n_rows)
        table = Table(displayName="SafeFiles", ref=f"A1:I{n_rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    autosize(ws, [70, 60, 32, 20, 10, 14, 12, 14, 8])

    # ---- Sheet: summary (first/active sheet) ----
    ws2 = wb["Summary"]
    disk = data["disk"]
    r = 1

    def line(a, b=""):
        nonlocal r
        ws2.cell(row=r, column=1, value=a)
        ws2.cell(row=r, column=2, value=b)
        r += 1

    ws2.cell(row=r, column=1, value="Storage Summary").font = Font(bold=True, size=14)
    r += 2
    line("Scanned at", data["scanned_at"])
    line("Drive", data["root"])
    line("Total capacity", human_size(disk["total"]))
    line("Used", human_size(disk["used"]))
    line("Free", human_size(disk["free"]))
    line("Files scanned", f"{data['file_count']:,}")
    line("Folders scanned", f"{data['dir_count']:,}")
    line("Skipped (no permission)", f"{data['inaccessible']:,}")
    r += 1
    line("Safe-to-review data (your files)", human_size(data["safe_total"]))
    line("System / application-critical data (excluded from the file list)", human_size(data["critical_total"]))
    r += 1
    line(f"Files listed in 'Safe to Delete - Files' (>= 1 MB)", f"{data['safe_files_count']:,}")
    if data.get("safe_files_truncated"):
        line("Note", "List capped at the largest 5,000 qualifying files.")
    r += 2

    ws2.cell(row=r, column=1, value="By file type (all scanned files)").font = Font(bold=True)
    r += 1
    ws2.cell(row=r, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=r, column=2, value="Size").font = Font(bold=True)
    r += 1
    for cat, size in sorted(data["category_sizes"].items(), key=lambda kv: -kv[1]):
        line(cat, human_size(size))
    r += 1

    ws2.cell(row=r, column=1, value="By top-level folder").font = Font(bold=True)
    r += 1
    ws2.cell(row=r, column=1, value="Folder").font = Font(bold=True)
    ws2.cell(row=r, column=2, value="Size").font = Font(bold=True)
    r += 1
    for item in data["top_level_folders"]:
        line(item["path"], human_size(item["size"]))
    r += 1

    ws2.cell(row=r, column=1, value="Junk / cache locations found").font = Font(bold=True)
    r += 1
    ws2.cell(row=r, column=1, value="Location").font = Font(bold=True)
    ws2.cell(row=r, column=2, value="Size").font = Font(bold=True)
    ws2.cell(row=r, column=3, value="Path").font = Font(bold=True)
    r += 1
    for label, info in sorted(data["junk"].items(), key=lambda kv: -kv[1]["size"]):
        if info["exists"]:
            ws2.cell(row=r, column=1, value=label)
            ws2.cell(row=r, column=2, value=human_size(info["size"]))
            ws2.cell(row=r, column=3, value=info["path"])
            r += 1

    autosize(ws2, [55, 20, 60])

    wb.save(out_path)


def build_folder_excel(folder_path, files, out_path):
    """A file-wise export scoped to one folder (and its subfolders) - `files`
    is already filtered by the caller to paths under `folder_path`."""
    wb = Workbook()
    wb.active.title = "Summary"

    ws = wb.create_sheet("Folder Files")
    headers = ["Path", "Folder", "File name", "Category", "Extension", "Size (bytes)", "Size", "Action", "Open"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    files_sorted = sorted(files, key=lambda f: -f["size"])
    total_size = 0
    for f in files_sorted:
        path = f["path"]
        folder, name, ext = _split_path(path)
        total_size += f["size"]
        row = [path, folder, name, f.get("category", ""), ext, f["size"], human_size(f["size"]), "", "Open"]
        ws.append(row)
        r = ws.max_row
        open_cell = ws.cell(row=r, column=9)
        open_cell.hyperlink = path
        open_cell.font = HYPERLINK_FONT
        open_cell.alignment = Alignment(horizontal="center")

    n_rows = len(files_sorted) + 1
    add_action_options_sheet(wb)
    if n_rows > 1:
        add_action_dropdown(ws, "H", n_rows)
        table = Table(displayName="FolderFiles", ref=f"A1:I{n_rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    autosize(ws, [70, 60, 32, 20, 10, 14, 12, 14, 8])

    ws2 = wb["Summary"]
    ws2.cell(row=1, column=1, value="Folder Export").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Folder")
    ws2.cell(row=3, column=2, value=folder_path)
    ws2.cell(row=4, column=1, value="Files listed (>= 256 KB)")
    ws2.cell(row=4, column=2, value=f"{len(files_sorted):,}")
    ws2.cell(row=5, column=1, value="Total size")
    ws2.cell(row=5, column=2, value=human_size(total_size))
    autosize(ws2, [28, 60])

    wb.save(out_path)


CLEANUP_SECTIONS = [
    ("temp_cache_files", "Temp/Cache"),
    ("old_downloads", "Old Download"),
    ("large_files", "Large File"),
    ("unknown_files", "Unknown"),
]


def build_cleanup_excel(report, out_path):
    """Export of everything flagged in the Cleanup Center: every duplicate
    extra copy plus temp/cache, old downloads, large files, and unknown-type
    items, each with its risk/confidence/reason and an Action dropdown."""
    wb = Workbook()
    wb.active.title = "Summary"

    ws = wb.create_sheet("Cleanup Items")
    headers = ["Category", "Path", "File name", "Extension", "Size (bytes)", "Size",
               "Risk", "Confidence", "Reason", "Action", "Open"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    def add_row(category, item, risk, confidence, reason):
        path = item["path"]
        _, name, ext = _split_path(path)
        ws.append([category, path, name, ext, item["size"], human_size(item["size"]),
                   risk, confidence, reason, "", "Open"])
        r = ws.max_row
        open_cell = ws.cell(row=r, column=11)
        open_cell.hyperlink = path
        open_cell.font = HYPERLINK_FONT
        open_cell.alignment = Alignment(horizontal="center")

    for group in report["duplicates"]["groups"]:
        for extra in group["extras"]:
            add_row("Duplicate (extra copy)", extra, "Low", 95, extra["reason"])
    for key, label in CLEANUP_SECTIONS:
        for item in report.get(key, []):
            add_row(label, item, item["risk_level"].capitalize(), item["confidence"], item["reason"])

    n_rows = ws.max_row
    add_action_options_sheet(wb)
    if n_rows > 1:
        add_action_dropdown(ws, "J", n_rows)
        table = Table(displayName="CleanupItems", ref=f"A1:K{n_rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    autosize(ws, [20, 65, 32, 10, 14, 12, 10, 12, 50, 14, 8])

    ws2 = wb["Summary"]
    r = 1

    def line(a, b=""):
        nonlocal r
        ws2.cell(row=r, column=1, value=a)
        ws2.cell(row=r, column=2, value=b)
        r += 1

    ws2.cell(row=r, column=1, value="Cleanup Center Summary").font = Font(bold=True, size=14)
    r += 2
    line("Generated at", report["generated_at"])
    line("Estimated recoverable space", human_size(report["recoverable_estimate_bytes"]))
    line("Duplicate waste", human_size(report["duplicates"]["total_waste_bytes"]))
    r += 1
    counts = report["counts"]
    ws2.cell(row=r, column=1, value="By category").font = Font(bold=True)
    r += 1
    line("Duplicates (extra copies)", f"{counts['duplicates']:,}")
    line("Temp/Cache files", f"{counts['temp_cache']:,}")
    line("Old downloads", f"{counts['old_downloads']:,}")
    line("Large files", f"{counts['large_files']:,}")
    line("Unknown (needs review)", f"{counts['unknown']:,}")
    autosize(ws2, [40, 20])

    wb.save(out_path)
